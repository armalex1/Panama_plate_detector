from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
import shutil
import cv2
import numpy as np
import easyocr
from processing import detect_plates, enhance_plate_image
from ocr import read_plate_text, extract_valid_plates
from verifier import load_plates_from_excel, check_plate
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

reader = easyocr.Reader(['en'])
UPLOAD_PATH = Path("uploaded_plates.xlsx")
TEMPLATE_PATH = Path("template.xlsx")
LOG_PATH = Path("placas_detectadas_log.xlsx")


def log_plate_detection(
    plate, authorized, rect_count, match_type, lat, lon, fecha, hora
):
    if LOG_PATH.exists():
        wb = load_workbook(LOG_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Fecha",
            "Hora",
            "Placa",
            "Autorizada",
            "Rectángulos detectados",
            "Tipo coincidencia",
            "Latitud",
            "Longitud"
        ])
    ws.append([
        fecha,
        hora,
        plate,
        "Sí" if authorized else "No",
        rect_count,
        match_type,
        lat,
        lon
    ])
    wb.save(LOG_PATH)


def load_active_plate_list():
    if UPLOAD_PATH.exists():
        return load_plates_from_excel(str(UPLOAD_PATH))
    return load_plates_from_excel("placas.xlsx")


@app.get("/", response_class=HTMLResponse)
def index():
    if LOG_PATH.exists():
        LOG_PATH.unlink()
    with open("static/index.html", "r", encoding="utf-8") as f:
        return f.read()


@app.post("/detect")
async def detect(
    file: UploadFile = File(...),
    lat: float = 37.3891,
    lon: float = -5.9845
):
    contents = await file.read()
    npimg = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)

    print("\U0001F4F8 Imagen recibida")

    results = []
    rects = detect_plates(img)
    print(f"\U0001F50D Rectángulos detectados: {len(rects)}")

    plate_list = load_active_plate_list()

    now = datetime.now()
    fecha = now.strftime("%Y-%m-%d")
    hora = now.strftime("%H:%M:%S")

    for x, y, w, h in rects:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        cropped = img[y: y + h, x: x + w]
        enhanced = enhance_plate_image(cropped)
        raw_text = read_plate_text(reader, enhanced)
        plates_detected = extract_valid_plates(raw_text)
        print(f"\U0001F9E0 OCR Result: {plates_detected}")

        for plate in plates_detected:
            if check_plate(plate, plate_list):
                found = True
                match_type = "exact"
            else:
                found = False
                match_type = (
                    "partial"
                    if any(
                        plate in p or p in plate
                        for p in plate_list
                    )
                    else "none"
                )
            log_plate_detection(
                plate, found, len(rects), match_type, lat, lon, fecha, hora
            )
            results.append({
                "plate": plate,
                "authorized": found,
                "lat": lat,
                "lon": lon,
                "fecha": fecha,
                "hora": hora,
                "match_type": match_type
            })

    cv2.imwrite("static/last_detected.jpg", img)
    return {"results": results, "rects": rects}


@app.post("/upload-plates")
async def upload_plates(file: UploadFile = File(...)):
    try:
        with open(UPLOAD_PATH, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}


@app.get("/template")
async def get_template():
    return FileResponse(
        TEMPLATE_PATH,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        filename="template.xlsx"
    )


@app.get("/download-log")
def download_log():
    if not LOG_PATH.exists():
        return {"error": "Log file not found."}
    return FileResponse(
        LOG_PATH,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        filename="placas_detectadas_log.xlsx"
    )
